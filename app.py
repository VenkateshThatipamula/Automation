import openpyxl as xl
from openpyxl.chart import BarChart, Reference


w_book=xl.load_workbook('mobileprices.xlsx')
sheet=w_book['Sheet1']
for row in range(3,sheet.max_row+1):
    if range(3,19):
        mobile_price = sheet.cell(row, 6)
        disc_mobile_price =mobile_price.value * 0.8
        disc_mobile_price_cell=sheet.cell(row,7)
        disc_mobile_price_cell.value=disc_mobile_price
    elif range(19,39):
        mobile_price = sheet.cell(row, 6)
        disc_mobile_price = mobile_price.value * 0.85
        disc_mobile_price_cell = sheet.cell(row, 7)
        disc_mobile_price_cell.value = disc_mobile_price
    else:
        mobile_price = sheet.cell(row, 6)
        disc_mobile_price = mobile_price.value * 0.9
        disc_mobile_price_cell = sheet.cell(row, 7)
        disc_mobile_price_cell.value = disc_mobile_price


values=Reference(sheet,min_row=3, max_row=sheet.max_row,min_col=6, max_col=7)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'i3')
w_book.save('mobleprices4.xlsx')


